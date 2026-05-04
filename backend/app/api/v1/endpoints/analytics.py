import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.dependencies import get_current_tenant
from app.models.tenant import Tenant
from app.schemas.analytics import (
    AnalyticsOverviewResponse,
    CategoryMetricRow,
    MonthlyFlowRow,
    ProviderMetricRow,
    TaxMonthlyFlowRow,
    TopThirdPartyRow,
)
from app.services.analytics_service import AnalyticsService

router = APIRouter(prefix="/analytics", tags=["Analytics"])


@router.get("/overview", response_model=AnalyticsOverviewResponse)
def get_analytics_overview(
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date] = Query(default=None),
    current_tenant: Tenant = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    return AnalyticsService.get_overview(db, current_tenant.id, date_from=date_from, date_to=date_to)


@router.get("/monthly-flow", response_model=list[MonthlyFlowRow])
def get_monthly_flow(
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date] = Query(default=None),
    current_tenant: Tenant = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    return AnalyticsService.get_monthly_flow(db, current_tenant.id, date_from=date_from, date_to=date_to)


@router.get("/top-customers", response_model=list[TopThirdPartyRow])
def get_top_customers(
    limit: int = Query(default=5, ge=1, le=20),
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date] = Query(default=None),
    current_tenant: Tenant = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    return AnalyticsService.get_top_customers(db, current_tenant.id, limit, date_from=date_from, date_to=date_to)


@router.get("/top-suppliers", response_model=list[TopThirdPartyRow])
def get_top_suppliers(
    limit: int = Query(default=5, ge=1, le=20),
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date] = Query(default=None),
    current_tenant: Tenant = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    return AnalyticsService.get_top_suppliers(db, current_tenant.id, limit, date_from=date_from, date_to=date_to)


@router.get("/expenses-by-category", response_model=list[CategoryMetricRow])
def get_expenses_by_category(
    limit: int = Query(default=6, ge=1, le=20),
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date] = Query(default=None),
    current_tenant: Tenant = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    return AnalyticsService.get_expenses_by_category(db, current_tenant.id, limit, date_from=date_from, date_to=date_to)


@router.get("/income-by-category", response_model=list[CategoryMetricRow])
def get_income_by_category(
    limit: int = Query(default=6, ge=1, le=20),
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date] = Query(default=None),
    current_tenant: Tenant = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    return AnalyticsService.get_income_by_category(db, current_tenant.id, limit, date_from=date_from, date_to=date_to)


@router.get("/tax-monthly-flow", response_model=list[TaxMonthlyFlowRow])
def get_tax_monthly_flow(
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date] = Query(default=None),
    current_tenant: Tenant = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    return AnalyticsService.get_tax_monthly_flow(db, current_tenant.id, date_from=date_from, date_to=date_to)


@router.get("/export")
def export_dashboard(
    date_from: Optional[date] = Query(default=None),
    date_to: Optional[date] = Query(default=None),
    current_tenant: Tenant = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    """Genera un Excel multi-hoja con toda la información del dashboard."""
    overview = AnalyticsService.get_overview(db, current_tenant.id, date_from=date_from, date_to=date_to)
    monthly = AnalyticsService.get_monthly_flow(db, current_tenant.id, date_from=date_from, date_to=date_to)
    suppliers = AnalyticsService.get_top_suppliers(db, current_tenant.id, 20, date_from=date_from, date_to=date_to)
    customers = AnalyticsService.get_top_customers(db, current_tenant.id, 20, date_from=date_from, date_to=date_to)
    exp_cat = AnalyticsService.get_expenses_by_category(db, current_tenant.id, 20, date_from=date_from, date_to=date_to)
    inc_cat = AnalyticsService.get_income_by_category(db, current_tenant.id, 20, date_from=date_from, date_to=date_to)
    tax = AnalyticsService.get_tax_monthly_flow(db, current_tenant.id, date_from=date_from, date_to=date_to)

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    def style_header_row(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    kpi_pairs = [
        ("Total ingresos (€)", overview.get("total_income")),
        ("Total gastos (€)", overview.get("total_expenses")),
        ("Beneficio neto (€)", overview.get("net_profit")),
        ("IVA repercutido (€)", overview.get("vat_charged")),
        ("IVA soportado (€)", overview.get("vat_supported")),
        ("Balance IVA (€)", overview.get("vat_balance")),
        ("Retenciones recibidas (€)", overview.get("retention_sales")),
        ("Retenciones pagadas renta (€)", overview.get("retention_rent")),
        ("Gasto fijo mensual (€)", overview.get("fixed_burn_rate")),
        ("Gasto variable mensual (€)", overview.get("variable_burn_rate")),
    ]
    ws.cell(row=1, column=1, value="KPI").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=2, value="Valor").fill = header_fill
    ws.cell(row=1, column=2).font = header_font
    for r, (label, value) in enumerate(kpi_pairs, 2):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=float(value or 0))
    autofit(ws)

    # ── Hoja 2: Flujo mensual ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Flujo mensual")
    style_header_row(ws2, ["Mes", "Ingresos (€)", "Gastos (€)", "Beneficio (€)"])
    for r, row in enumerate(monthly, 2):
        inc = float(row.get("income") or 0)
        exp = float(row.get("expenses") or 0)
        ws2.cell(row=r, column=1, value=row.get("month"))
        ws2.cell(row=r, column=2, value=inc)
        ws2.cell(row=r, column=3, value=exp)
        ws2.cell(row=r, column=4, value=inc - exp)
    autofit(ws2)

    # ── Hoja 3: IVA mensual ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("IVA mensual")
    style_header_row(ws3, ["Mes", "IVA Repercutido (€)", "IVA Soportado (€)", "Balance IVA (€)"])
    for r, row in enumerate(tax, 2):
        charged = float(row.get("vat_charged") or 0)
        supported = float(row.get("vat_supported") or 0)
        ws3.cell(row=r, column=1, value=row.get("month"))
        ws3.cell(row=r, column=2, value=charged)
        ws3.cell(row=r, column=3, value=supported)
        ws3.cell(row=r, column=4, value=charged - supported)
    autofit(ws3)

    # ── Hoja 4: Top proveedores ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Top Proveedores")
    style_header_row(ws4, ["Proveedor", "Total (€)"])
    for r, row in enumerate(suppliers, 2):
        ws4.cell(row=r, column=1, value=row.get("name"))
        ws4.cell(row=r, column=2, value=float(row.get("amount") or 0))
    autofit(ws4)

    # ── Hoja 5: Top clientes ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("Top Clientes")
    style_header_row(ws5, ["Cliente", "Total (€)"])
    for r, row in enumerate(customers, 2):
        ws5.cell(row=r, column=1, value=row.get("name"))
        ws5.cell(row=r, column=2, value=float(row.get("amount") or 0))
    autofit(ws5)

    # ── Hoja 6: Gastos por categoría ───────────────────────────────────────────
    ws6 = wb.create_sheet("Gastos por categoría")
    style_header_row(ws6, ["Categoría", "Total (€)"])
    for r, row in enumerate(exp_cat, 2):
        ws6.cell(row=r, column=1, value=row.get("category_name"))
        ws6.cell(row=r, column=2, value=float(row.get("total_amount") or 0))
    autofit(ws6)

    # ── Hoja 7: Ingresos por categoría ────────────────────────────────────────
    ws7 = wb.create_sheet("Ingresos por categoría")
    style_header_row(ws7, ["Categoría", "Total (€)"])
    for r, row in enumerate(inc_cat, 2):
        ws7.cell(row=r, column=1, value=row.get("category_name"))
        ws7.cell(row=r, column=2, value=float(row.get("total_amount") or 0))
    autofit(ws7)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    period_label = ""
    if date_from and date_to:
        period_label = f"_{date_from}_{date_to}"
    elif date_from:
        period_label = f"_desde_{date_from}"

    filename = f"dashboard_{current_tenant.name}{period_label}_{date.today()}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
